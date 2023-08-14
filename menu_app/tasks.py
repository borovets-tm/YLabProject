"""Модуль фоновых задач, обрабатываемых Celery."""
import csv
from io import StringIO
from uuid import UUID

import requests
from openpyxl import Workbook, load_workbook
from requests import Response
from sqlalchemy import delete

from menu_app.config import config
from menu_app.models import Dish, Menu, Submenu
from menu_app.worker import SyncSessionLocal, celery

fields_menu = ('id', 'title', 'description')
fields_submenu = ('menu_id', 'id', 'title', 'description')
fields_dish = (
    'submenu_id',
    'id',
    'title',
    'description',
    'price',
    'discount'
)


def _get_google_sheets_data() -> list:
    """
    Функция читает документ Google Sheets в публичном доступе и возвращает\
    данные в виде списка для дальнейшего выполнения фоновой задачи.

    :return: Список данных для обработки.
    """
    r: Response = requests.get(
        'https://docs.google.com/spreadsheets/d/e/2PACX-1vR6a115vW2IsS9ebkRNK3'
        'S_GqCkVieVPxbxjTjNLWq5n99jFw0yY9W0BWhvSHqsIRSEbktRKTKIGBz1/pub?gid=0'
        '&single=true&output=csv'
    )

    result: StringIO = StringIO(r.content.decode())
    reader = csv.reader(result, delimiter=',')
    menu_data_list: list = []
    submenu_data_list: list = []
    dish_data_list: list = []
    menu_id: UUID | None = None
    submenu_id: UUID | None = None
    data_list: list = [';'.join(row).rstrip(';').split(';') for row in reader]
    for data in data_list:
        if len(data) == 3:
            menu_id = UUID(data[0])
            data[0] = menu_id
            menu_data_list.append(data)
        elif len(data) == 4:
            data[0] = menu_id
            submenu_id = UUID(data[1])
            data[1] = submenu_id
            submenu_data_list.append(data)
        else:
            data[1] = submenu_id
            data[2] = UUID(data[2])
            if len(data) < 7:
                data.append(0)
            dish_data_list.append(data[1:])
    return [menu_data_list, submenu_data_list, dish_data_list]


def _get_file_data(wb: Workbook) -> list:
    """
    Функция читает файл Menu.xlsx и возвращает полученные данные для \
    дальнейшего выполнения фоновой задачи.

    :param wb: Файл книги excel.
    :return: Список данных для обработки.
    """
    sheet = wb.active
    data_list: list = []
    menu_id: UUID | None = None
    submenu_id: UUID | None = None
    for i in range(1, sheet.max_row + 1):
        data_entity = []
        for j in range(1, sheet.max_column + 1):
            cell_obj = sheet.cell(row=i, column=j).value
            if cell_obj:
                if j in [1, 2, 3]:
                    try:
                        cell_obj = UUID(cell_obj)
                        if j == 1:
                            menu_id = cell_obj
                        elif j == 2:
                            submenu_id = cell_obj
                    except ValueError:
                        pass
            else:
                if j == 1:
                    cell_obj = menu_id
                elif j == 2:
                    cell_obj = submenu_id
                else:
                    continue
            data_entity.append(cell_obj)
        data_list.append(data_entity)
    menu_data_list = list(filter(lambda x: len(x) == 3, data_list))
    submenu_data_list = list(filter(lambda x: len(x) == 4, data_list))
    dish_data_list = list(
        map(
            lambda x: x[1:] + [0], filter(lambda x: len(x) > 4, data_list)
        )
    )
    return [menu_data_list, submenu_data_list, dish_data_list]


def _get_file() -> Workbook:
    """
    Функция загружает книгу excel по указанному пути.

    :return: Файл книги excel.
    """
    return load_workbook(config.BASE_DIR / 'admin/Menu.xlsx')


@celery.task
def update_db_from_excel() -> None:
    """
    Функция обновляет данные в базе из excel файла в фоновом режиме каждые 15\
     секунд. А если файл недоступен обновление производится из Google Sheets.

    :return: None.
    """
    id_dict: dict = {
        'menu': [],
        'submenu': [],
        'dish': []
    }
    try:
        wb: Workbook = _get_file()
        data_list = _get_file_data(wb)
    except FileNotFoundError:
        data_list = _get_google_sheets_data()
    menu_data_list = data_list[0]
    submenu_data_list = data_list[1]
    dish_data_list = data_list[2]
    id_dict['menu'].extend([data[0] for data in menu_data_list])
    id_dict['submenu'].extend([data[1] for data in submenu_data_list])
    id_dict['dish'].extend([data[1] for data in dish_data_list])
    menu_list = [
        Menu(**dict(zip(fields_menu, data))) for data in menu_data_list
    ]
    submenu_list = [
        Submenu(**dict(zip(fields_submenu, data)))
        for data in submenu_data_list
    ]
    dish_list = [
        Dish(**dict(zip(fields_dish, data))) for data in dish_data_list
    ]
    obj_list = menu_list + submenu_list + dish_list
    with SyncSessionLocal() as db:
        for obj in obj_list:
            db.merge(obj)
        query_menu = delete(Menu).filter(Menu.id.not_in(id_dict['menu']))
        query_submenu = delete(Submenu).filter(
            Submenu.id.not_in(id_dict['submenu'])
        )
        query_dish = delete(Dish).filter(Dish.id.not_in(id_dict['dish']))
        db.execute(query_menu)
        db.execute(query_submenu)
        db.execute(query_dish)
        db.commit()


if __name__ == '__main__':
    update_db_from_excel()
