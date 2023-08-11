"""Сервисный слой приложения, не связанного с конкретной моделью приложения."""
from uuid import UUID

from openpyxl import load_workbook
from sqlalchemy import Sequence
from sqlalchemy.ext.asyncio import AsyncSession

from menu_app.models import Dish, Menu, Submenu
from menu_app.repositories.app_repository import get_tree_menu_repository, merge_objects
from menu_app.services.base_service import BaseService


class AppService(BaseService):
    """Класс сервисных методов приложения, не связанных с конкретной \
    моделью."""

    async def get_full_menu(self, db: AsyncSession) -> Sequence:
        """
        Метод обрабатывает запрос на получения всех данных из БД в виде \
        дерева JSON. Проверяет наличие кэша и, при его отсутствии, запишет кэш.

        :param db: Экземпляр сеанса базы данных.
        :return: Древовидное меню со всеми элементами БД.
        """
        cache = await self.get_cache(self.full_menu)
        if cache:
            return cache
        result = await get_tree_menu_repository(db)
        await self.set_cache(self.full_menu, result)
        return result

    @classmethod
    async def read_excel(cls) -> list:
        """
        Метод читает excel файл с данными из папки admin проекта и формирует \
        список с полученными данными.

        :return: Список со списком полей моделей.
        """
        wb = load_workbook('menu_app/admin/Menu.xlsx')
        sheet = wb.active
        read_data = []
        menu_id = None
        submenu_id = None
        for i in range(1, sheet.max_row + 1):
            data_entity = []
            for j in range(1, sheet.max_column + 1):
                cell_obj = sheet.cell(row=i, column=j).value
                if cell_obj:
                    if j in [1, 2, 3]:
                        try:
                            cell_obj = UUID(cell_obj)
                            if j == 1:
                                menu_id = cell_obj
                            elif j == 2:
                                submenu_id = cell_obj
                        except ValueError:
                            pass
                else:
                    if j == 1:
                        cell_obj = menu_id
                    elif j == 2:
                        cell_obj = submenu_id
                    else:
                        continue
                data_entity.append(cell_obj)
            read_data.append(data_entity)
        return read_data

    async def preparing_data_for_updating(self, db: AsyncSession) -> None:
        """
        Метод подготавливает объекты для обновления или добавления в БД, \
        формируя список объектов из данных полученных после парсинга excel.

        :param db: Экземпляр сеанса базы данных.
        :return: None.
        """
        fields_menu = ('id', 'title', 'description')
        fields_submenu = ('menu_id', 'id', 'title', 'description')
        fields_dish = ('submenu_id', 'id', 'title', 'description', 'price')
        data_list = await self.read_excel()
        menu_data_list = list(filter(lambda x: len(x) == 3, data_list))
        submenu_data_list = list(filter(lambda x: len(x) == 4, data_list))
        dish_data_list = list(filter(lambda x: len(x) > 4, data_list))

        menu_data_list = [
            Menu(**dict(zip(fields_menu, data))) for data in menu_data_list
        ]
        submenu_data_list = [
            Submenu(**dict(zip(fields_submenu, data)))
            for data in submenu_data_list
        ]
        dish_data_list = [
            Dish(**dict(zip(fields_dish, data[1:]))) for data in dish_data_list
        ]
        all_data_list = menu_data_list + submenu_data_list + dish_data_list
        await merge_objects(db, all_data_list)


service: AppService = AppService()
